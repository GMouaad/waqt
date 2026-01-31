"""Utility functions for time tracking calculations."""

import calendar
import csv
import io
import json
from datetime import datetime, timedelta, date, time as datetime_time
from typing import List, Dict, Tuple, Optional
import openpyxl
from openpyxl.styles import Font
from .models import TimeEntry, LeaveDay, Settings


def is_weekend(check_date: date) -> bool:
    """
    Check if a given date falls on a weekend (Saturday or Sunday).

    Args:
        check_date: Date to check

    Returns:
        True if the date is Saturday (5) or Sunday (6), False otherwise
    """
    return check_date.weekday() in (5, 6)


def get_date_range(start_date: date, end_date: date) -> List[date]:
    """
    Generate a list of dates between start_date and end_date (inclusive).

    Args:
        start_date: Start date (inclusive)
        end_date: End date (inclusive)

    Returns:
        List of date objects from start_date to end_date
    """
    if start_date > end_date:
        return []

    date_list = []
    current_date = start_date
    while current_date <= end_date:
        date_list.append(current_date)
        current_date += timedelta(days=1)

    return date_list


def get_working_days_in_range(start_date: date, end_date: date) -> List[date]:
    """
    Get all working days (excluding weekends) in a date range.

    Args:
        start_date: Start date (inclusive)
        end_date: End date (inclusive)

    Returns:
        List of date objects that are working days (Monday-Friday)
    """
    all_dates = get_date_range(start_date, end_date)
    return [d for d in all_dates if not is_weekend(d)]


def calculate_leave_hours(start_date: date, end_date: date) -> Dict:
    """
    Calculate leave statistics for a date range.

    Args:
        start_date: Start date of leave
        end_date: End date of leave

    Returns:
        Dictionary with:
        - total_days: Total calendar days in range
        - working_days: Number of working days (excluding weekends)
        - weekend_days: Number of weekend days
        - working_hours: Total working hours (working_days * standard_hours_per_day)
    """
    # Generate date range once and reuse it
    all_dates = get_date_range(start_date, end_date)
    working_days_list = [d for d in all_dates if not is_weekend(d)]

    total_days = len(all_dates)
    working_days = len(working_days_list)
    weekend_days = total_days - working_days

    # Get standard hours per day from settings
    standard_hours_per_day = get_standard_hours_per_day()
    working_hours = working_days * standard_hours_per_day

    return {
        "total_days": total_days,
        "working_days": working_days,
        "weekend_days": weekend_days,
        "working_hours": working_hours,
    }


def format_time(
    time_obj: Optional[datetime_time], time_format: Optional[str] = None
) -> str:
    """
    Format a time object according to the user's preferred time format.

    Args:
        time_obj: datetime.time object to format. If None, returns an empty string.
        time_format: Time format preference ('12' or '24'). If None, reads from settings.

    Returns:
        Formatted time string (e.g., '13:30' or '01:30 PM'), or empty string if time_obj is None
    """
    if time_obj is None:
        return ""

    if time_format is None:
        time_format = Settings.get_setting("time_format", "24")

    if time_format == "12":
        # Convert to 12-hour format with AM/PM
        hour = time_obj.hour
        minute = time_obj.minute
        am_pm = "AM" if hour < 12 else "PM"

        # Convert hour to 12-hour format
        display_hour = hour % 12
        if display_hour == 0:
            display_hour = 12

        return f"{display_hour:d}:{minute:02d} {am_pm}"
    else:
        # Default to 24-hour format
        return time_obj.strftime("%H:%M")


def parse_time_input(time_str: str, time_format: str = "24") -> datetime_time:
    """
    Parse a time string based on the configured format.

    Args:
        time_str: Time string to parse (e.g., '14:30' or '02:30 PM')
        time_format: Expected format ('12' or '24')

    Returns:
        datetime.time object

    Raises:
        ValueError: If time string format is invalid
    """
    if time_format == "12":
        try:
            # Normalize input to uppercase to ensure AM/PM matching is robust
            return datetime.strptime(time_str.upper(), "%I:%M %p").time()
        except ValueError:
            # Fallback to 24-hour format if 12-hour parsing fails
            return datetime.strptime(time_str, "%H:%M").time()
    else:
        return datetime.strptime(time_str, "%H:%M").time()


def get_standard_hours_per_day() -> float:
    """Get the configured standard hours per day from settings."""
    return Settings.get_float("standard_hours_per_day", 8.0)


def get_standard_hours_per_week() -> float:
    """Get the configured standard hours per week from settings."""
    return Settings.get_float("weekly_hours", 40.0)


def calculate_daily_overtime(
    duration_hours: float, standard_hours: Optional[float] = None
) -> float:
    """
    Calculate overtime hours for a single day.

    Args:
        duration_hours: Total hours worked in the day
        standard_hours: Standard work hours per day (default: from settings or 8)

    Returns:
        Overtime hours (0 if no overtime)
    """
    if standard_hours is None:
        standard_hours = get_standard_hours_per_day()
    return max(0, duration_hours - standard_hours)


def calculate_duration(start_time: datetime.time, end_time: datetime.time) -> float:
    """
    Calculate duration in hours between two times.

    Args:
        start_time: Start time
        end_time: End time

    Returns:
        Duration in hours
    """
    # Convert times to datetime for calculation
    today = datetime.today().date()
    start = datetime.combine(today, start_time)
    end = datetime.combine(today, end_time)

    # Handle case where end time is before start time (crosses midnight)
    if end < start:
        end += timedelta(days=1)

    duration = end - start
    return duration.total_seconds() / 3600  # Convert to hours


def get_week_bounds(date: datetime.date) -> Tuple[datetime.date, datetime.date]:
    """
    Get the start (Monday) and end (Sunday) dates of the week containing the given date.

    Args:
        date: Any date in the week

    Returns:
        Tuple of (week_start, week_end)
    """
    # Find Monday of the week
    week_start = date - timedelta(days=date.weekday())
    # Find Sunday of the week
    week_end = week_start + timedelta(days=6)
    return week_start, week_end


def get_month_bounds(date: datetime.date) -> Tuple[datetime.date, datetime.date]:
    """
    Get the first and last dates of the month containing the given date.

    Args:
        date: Any date in the month

    Returns:
        Tuple of (month_start, month_end)
    """
    month_start = date.replace(day=1)
    # Get last day of month
    if date.month == 12:
        month_end = date.replace(day=31)
    else:
        next_month = date.replace(month=date.month + 1, day=1)
        month_end = next_month - timedelta(days=1)
    return month_start, month_end


def calculate_weekly_stats(
    entries: List[TimeEntry], standard_hours: Optional[float] = None
) -> Dict:
    """
    Calculate weekly statistics from time entries.

    Args:
        entries: List of TimeEntry objects for the week
        standard_hours: Standard work hours per week (default: from settings or 40)

    Returns:
        Dictionary with total_hours, overtime, and working_days
    """
    if standard_hours is None:
        standard_hours = get_standard_hours_per_week()
    total_hours = sum(entry.duration_hours for entry in entries)
    overtime = max(0, total_hours - standard_hours)
    working_days = len(set(entry.date for entry in entries))

    return {
        "total_hours": round(total_hours, 2),
        "overtime": round(overtime, 2),
        "working_days": working_days,
        "standard_hours": standard_hours,
    }


def calculate_monthly_stats(
    entries: List[TimeEntry], leave_days: List[LeaveDay]
) -> Dict:
    """
    Calculate monthly statistics from time entries and leave days.

    Args:
        entries: List of TimeEntry objects for the month
        leave_days: List of LeaveDay objects for the month

    Returns:
        Dictionary with comprehensive monthly statistics
    """
    total_hours = sum(entry.duration_hours for entry in entries)
    working_days = len(set(entry.date for entry in entries))

    # Count leave days by type
    vacation_days = len(
        [leave for leave in leave_days if leave.leave_type == "vacation"]
    )
    sick_days = len([leave for leave in leave_days if leave.leave_type == "sick"])

    # Calculate expected hours using configured standard hours per day
    standard_hours_per_day = get_standard_hours_per_day()
    expected_hours = working_days * standard_hours_per_day
    overtime = max(0, total_hours - expected_hours)

    return {
        "total_hours": round(total_hours, 2),
        "overtime": round(overtime, 2),
        "working_days": working_days,
        "vacation_days": vacation_days,
        "sick_days": sick_days,
        "expected_hours": expected_hours,
    }


def format_hours(hours: float) -> str:
    """
    Format hours as HH:MM string.

    Args:
        hours: Hours as a float

    Returns:
        Formatted string like "8:30"
    """
    h = int(hours)
    m = int((hours - h) * 60)
    return f"{h}:{m:02d}"


def get_time_entries_for_period(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> List[TimeEntry]:
    """
    Get time entries for a specific date range.

    Note: If both start_date and end_date are None, all time entries
    in the database will be returned. For large datasets, this may
    impact performance. Consider providing a date range for better
    performance.

    Args:
        start_date: Optional start date (inclusive). If provided without
                   end_date, only start_date filtering is applied.
        end_date: Optional end date (inclusive). If provided without
                 start_date, only end_date filtering is applied.

    Returns:
        List of TimeEntry objects, ordered by date ascending
    """
    query = TimeEntry.query
    if start_date and end_date:
        query = query.filter(TimeEntry.date >= start_date, TimeEntry.date <= end_date)
    elif start_date:
        # Support filtering from start_date onwards
        query = query.filter(TimeEntry.date >= start_date)
    elif end_date:
        # Support filtering up to end_date
        query = query.filter(TimeEntry.date <= end_date)
    return query.order_by(TimeEntry.date.asc()).all()


def calculate_daily_overtime_for_entries(entries: List[TimeEntry]) -> Dict[date, float]:
    """
    Calculate daily overtime for a list of entries.

    Args:
        entries: List of TimeEntry objects

    Returns:
        Dictionary mapping date to overtime hours
    """
    daily_totals = {}
    for entry in entries:
        if entry.date not in daily_totals:
            daily_totals[entry.date] = 0.0
        daily_totals[entry.date] += entry.duration_hours

    daily_overtime = {}
    for date_key, total_hours in daily_totals.items():
        daily_overtime[date_key] = calculate_daily_overtime(total_hours)

    return daily_overtime


def export_time_entries_to_csv(
    entries: List[TimeEntry],
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> str:
    """
    Export time entries to CSV format.

    Args:
        entries: List of TimeEntry objects to export
        start_date: Optional start date for the export period
        end_date: Optional end date for the export period

    Returns:
        CSV content as a string
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    headers = [
        "Date",
        "Day of Week",
        "Start Time",
        "End Time",
        "Duration (Hours)",
        "Duration (HH:MM)",
        "Description",
        "Category",
        "Overtime",
        "Created At",
    ]
    writer.writerow(headers)

    # Calculate daily totals for overtime
    daily_overtime = calculate_daily_overtime_for_entries(entries)

    # Write data rows
    for entry in entries:
        # Get the overtime for this entry's day
        overtime = daily_overtime[entry.date]

        row = [
            entry.date.isoformat(),
            entry.date.strftime("%A"),
            format_time(entry.start_time),
            format_time(entry.end_time),
            f"{entry.duration_hours:.2f}",
            format_hours(entry.duration_hours),
            entry.description,
            entry.category.name if entry.category else "",
            f"{overtime:.2f}",
            entry.created_at.isoformat() if entry.created_at else "",
        ]
        writer.writerow(row)

    # Add summary statistics if there are entries
    if entries:
        writer.writerow([])  # Empty row
        writer.writerow(["Summary Statistics"])

        # Format period display
        if start_date and end_date:
            period_str = f"{start_date} to {end_date}"
        else:
            period_str = "All time entries"

        writer.writerow(["Period", period_str])
        writer.writerow(["Total Entries", len(entries)])
        total_hours = sum(entry.duration_hours for entry in entries)
        writer.writerow(["Total Hours", f"{total_hours:.2f}"])
        writer.writerow(["Total Hours (HH:MM)", format_hours(total_hours)])
        working_days = len(set(entry.date for entry in entries))
        writer.writerow(["Working Days", working_days])
        # Calculate total overtime from daily overtime values
        total_overtime = sum(daily_overtime.values())
        writer.writerow(["Total Overtime", f"{total_overtime:.2f}"])

    return output.getvalue()


def export_time_entries_to_json(
    entries: List[TimeEntry],
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    leave_days: Optional[List[LeaveDay]] = None,
) -> str:
    """
    Export time entries to JSON format.

    Args:
        entries: List of TimeEntry objects to export
        start_date: Optional start date for the export period
        end_date: Optional end date for the export period
        leave_days: Optional list of LeaveDay objects to include

    Returns:
        JSON content as a string
    """
    data = []

    # Calculate daily totals for overtime
    daily_overtime = calculate_daily_overtime_for_entries(entries)

    for entry in entries:
        overtime = daily_overtime[entry.date]
        entry_data = {
            "id": entry.id,
            "date": entry.date.isoformat(),
            "start_time": entry.start_time.strftime("%H:%M:%S"),
            "end_time": entry.end_time.strftime("%H:%M:%S"),
            "duration_hours": round(entry.duration_hours, 2),
            "description": entry.description,
            "overtime": round(overtime, 2),
            "category": entry.category.name if entry.category else None,
            "category_code": entry.category.code if entry.category else None,
            "created_at": entry.created_at.isoformat() if entry.created_at else None,
        }
        data.append(entry_data)

    # Build leave days data
    leave_data = []
    if leave_days:
        for leave in leave_days:
            leave_data.append(
                {
                    "id": leave.id,
                    "date": leave.date.isoformat(),
                    "leave_type": leave.leave_type,
                    "description": leave.description or "",
                    "created_at": (
                        leave.created_at.isoformat() if leave.created_at else None
                    ),
                }
            )

    export_data = {
        "period": {
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
        },
        "entries": data,
        "leave_days": leave_data,
        "summary": {
            "total_entries": len(entries),
            "total_hours": round(sum(e.duration_hours for e in entries), 2),
            "total_overtime": round(sum(daily_overtime.values()), 2),
            "total_leave_days": len(leave_data),
        },
    }

    return json.dumps(export_data, indent=2)


def export_time_entries_to_excel(
    entries: List[TimeEntry],
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> bytes:
    """
    Export time entries to Excel (XLSX) format.

    Args:
        entries: List of TimeEntry objects to export
        start_date: Optional start date for the export period
        end_date: Optional end date for the export period

    Returns:
        Excel file content as bytes
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Time Entries"

    # Headers
    headers = [
        "Date",
        "Day of Week",
        "Start Time",
        "End Time",
        "Duration (Hours)",
        "Duration (HH:MM)",
        "Description",
        "Category",
        "Overtime",
        "Created At",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Calculate daily totals for overtime
    daily_overtime = calculate_daily_overtime_for_entries(entries)

    # Data rows
    for row_idx, entry in enumerate(entries, 2):
        overtime = daily_overtime[entry.date]

        row_data = [
            entry.date,  # openpyxl handles dates
            entry.date.strftime("%A"),
            entry.start_time,  # openpyxl handles time
            entry.end_time,
            entry.duration_hours,
            format_hours(entry.duration_hours),
            entry.description,
            entry.category.name if entry.category else "",
            overtime,
            entry.created_at,
        ]

        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths using header and a limited number of sample rows
    max_sample_rows = 200
    last_row = sheet.max_row
    sample_last_row = min(last_row, max_sample_rows)

    for col_idx, header in enumerate(headers, 1):
        # Start with the header length
        max_length = len(str(header)) if header is not None else 0

        # Sample only up to max_sample_rows data rows for this column
        for row_idx in range(2, sample_last_row + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                cell_length = len(str(cell_value))
                if cell_length > max_length:
                    max_length = cell_length

        column_letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[column_letter].width = max_length + 2

    # Summary sheet
    summary_sheet = workbook.create_sheet(title="Summary")

    summary_data = [
        ("Period Start", start_date if start_date else "All time"),
        ("Period End", end_date if end_date else "All time"),
        ("Total Entries", len(entries)),
        ("Total Hours", sum(e.duration_hours for e in entries)),
        ("Total Overtime", sum(daily_overtime.values())),
        ("Working Days", len(set(e.date for e in entries))),
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        summary_sheet.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        summary_sheet.cell(row=row_idx, column=2, value=value)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def generate_calendar_data(year: int, month: int) -> Dict:
    """
    Generate calendar data for a given month with entry information.

    Args:
        year: Year for the calendar
        month: Month for the calendar (1-12)

    Returns:
        Dictionary containing:
            - weeks: List of weeks, each week is a list of day dicts
            - month_name: Name of the month
            - year: Year
            - prev_month: Dict with year and month for previous month
            - next_month: Dict with year and month for next month
    """

    # Get month boundaries
    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year, 12, 31)
    else:
        next_month_start = date(year, month + 1, 1)
        month_end = next_month_start - timedelta(days=1)

    # Get all time entries for the month
    time_entries = TimeEntry.query.filter(
        TimeEntry.date >= month_start, TimeEntry.date <= month_end
    ).all()

    # Get all leave days for the month
    leave_days = LeaveDay.query.filter(
        LeaveDay.date >= month_start, LeaveDay.date <= month_end
    ).all()

    # Create dictionaries for quick lookup
    entries_by_date = {}
    for entry in time_entries:
        if entry.date not in entries_by_date:
            entries_by_date[entry.date] = []
        entries_by_date[entry.date].append(entry)

    leaves_by_date = {}
    for leave in leave_days:
        leaves_by_date[leave.date] = leave

    # Generate calendar weeks
    cal = calendar.monthcalendar(year, month)
    weeks = []

    today = datetime.now().date()

    for week in cal:
        week_days = []
        for day_num in week:
            if day_num == 0:
                # Day from another month
                week_days.append(
                    {
                        "day": "",
                        "is_current_month": False,
                        "is_today": False,
                        "has_entry": False,
                        "has_leave": False,
                        "leave_type": None,
                        "total_hours": 0,
                        "entry_count": 0,
                        "date": None,
                    }
                )
            else:
                day_date = date(year, month, day_num)
                has_entries = day_date in entries_by_date
                has_leave = day_date in leaves_by_date

                total_hours = 0
                entry_count = 0
                if has_entries:
                    entry_count = len(entries_by_date[day_date])
                    total_hours = sum(
                        e.duration_hours for e in entries_by_date[day_date]
                    )

                leave_type = None
                if has_leave:
                    leave_type = leaves_by_date[day_date].leave_type

                week_days.append(
                    {
                        "day": day_num,
                        "is_current_month": True,
                        "is_today": day_date == today,
                        "has_entry": has_entries,
                        "has_leave": has_leave,
                        "leave_type": leave_type,
                        "total_hours": round(total_hours, 2),
                        "entry_count": entry_count,
                        "date": day_date.isoformat(),
                    }
                )

        weeks.append(week_days)

    # Calculate previous and next month
    if month == 1:
        prev_month = {"year": year - 1, "month": 12}
    else:
        prev_month = {"year": year, "month": month - 1}

    if month == 12:
        next_month = {"year": year + 1, "month": 1}
    else:
        next_month = {"year": year, "month": month + 1}

    return {
        "weeks": weeks,
        "month_name": calendar.month_name[month],
        "year": year,
        "month": month,
        "prev_month": prev_month,
        "next_month": next_month,
    }


# =============================================================================
# Import Parsing Functions
# =============================================================================


def detect_import_format(file_path: str) -> str:
    """
    Auto-detect file format from extension.

    Args:
        file_path: Path to the file

    Returns:
        Format string: 'csv', 'json', or 'excel'

    Raises:
        ValueError: If format cannot be determined
    """
    import os

    ext = os.path.splitext(file_path)[1].lower()

    format_map = {
        ".csv": "csv",
        ".json": "json",
        ".xlsx": "excel",
        ".xls": "excel",
    }

    if ext in format_map:
        return format_map[ext]

    raise ValueError(
        f"Cannot determine file format from extension '{ext}'. "
        "Supported extensions: .csv, .json, .xlsx"
    )


def normalize_time_string(time_str: str) -> Optional[datetime_time]:
    """
    Parse various time formats into datetime.time.

    Supported formats:
    - HH:MM (24-hour)
    - HH:MM:SS (24-hour with seconds)
    - HH:MM AM/PM (12-hour)
    - HH:MM:SS AM/PM (12-hour with seconds)

    Args:
        time_str: Time string to parse

    Returns:
        datetime.time object or None if parsing fails
    """
    if not time_str or not isinstance(time_str, str):
        return None

    time_str = time_str.strip()

    # List of formats to try
    formats = [
        "%H:%M:%S",  # 14:30:00
        "%H:%M",  # 14:30
        "%I:%M:%S %p",  # 02:30:00 PM
        "%I:%M %p",  # 02:30 PM
        "%I:%M:%S%p",  # 02:30:00PM (no space)
        "%I:%M%p",  # 02:30PM (no space)
    ]

    for fmt in formats:
        try:
            parsed = datetime.strptime(time_str, fmt)
            return parsed.time()
        except ValueError:
            continue

    return None


def normalize_date_string(date_str: str) -> Optional[date]:
    """
    Parse various date formats into datetime.date.

    Supported formats (in order of precedence):
    - YYYY-MM-DD (ISO 8601 - preferred, unambiguous)
    - YYYY/MM/DD (ISO 8601 variant - unambiguous)
    - DD-MM-YYYY (European, unambiguous when day > 12)
    - DD/MM/YYYY (European)
    - MM/DD/YYYY (US format - tried last to prefer European interpretation)

    Note: For ambiguous dates like 01/02/2026, DD/MM/YYYY (European) is tried
    before MM/DD/YYYY (US). Users importing from US-formatted systems should
    ensure their dates are in ISO format (YYYY-MM-DD) to avoid misinterpretation.

    Args:
        date_str: Date string to parse

    Returns:
        datetime.date object or None if parsing fails
    """
    if not date_str or not isinstance(date_str, str):
        return None

    date_str = date_str.strip()

    # Formats listed in order of precedence
    # ISO formats first (unambiguous), then European, then US
    formats = [
        "%Y-%m-%d",  # 2026-01-15 (ISO 8601 - preferred)
        "%Y/%m/%d",  # 2026/01/15 (ISO variant)
        "%d-%m-%Y",  # 15-01-2026 (European with dashes)
        "%d/%m/%Y",  # 15/01/2026 (European with slashes)
        "%m/%d/%Y",  # 01/15/2026 (US format - tried last)
    ]

    for fmt in formats:
        try:
            parsed = datetime.strptime(date_str, fmt)
            return parsed.date()
        except ValueError:
            continue

    return None


def validate_time_entry_data(entry: Dict) -> Tuple[bool, List[str]]:
    """
    Validate a single time entry dictionary.

    Args:
        entry: Dictionary with time entry data

    Returns:
        Tuple of (is_valid, list_of_error_messages)
    """
    errors = []

    # Required fields
    if not entry.get("date"):
        errors.append("Missing required field: date")
    else:
        parsed_date = normalize_date_string(str(entry["date"]))
        if parsed_date is None:
            errors.append(f"Invalid date format: {entry['date']}")

    if not entry.get("start_time"):
        errors.append("Missing required field: start_time")
    else:
        parsed_start = normalize_time_string(str(entry["start_time"]))
        if parsed_start is None:
            errors.append(f"Invalid start_time format: {entry['start_time']}")

    if not entry.get("end_time"):
        errors.append("Missing required field: end_time")
    else:
        parsed_end = normalize_time_string(str(entry["end_time"]))
        if parsed_end is None:
            errors.append(f"Invalid end_time format: {entry['end_time']}")

    # Duration validation (warning, not error)
    if entry.get("duration_hours"):
        try:
            duration = float(entry["duration_hours"])
            if duration < 0:
                errors.append(f"Invalid negative duration: {duration}")
            elif duration > 24:
                # Warning but not an error - will be in warnings list
                pass
        except (ValueError, TypeError):
            pass  # Will be recalculated

    return len(errors) == 0, errors


def parse_time_entries_from_json(content: str) -> Dict:
    """
    Parse JSON content into structured import data.

    Args:
        content: JSON string content

    Returns:
        Dictionary with:
        - entries: List of entry dictionaries
        - leave_days: List of leave day dictionaries
        - period: Period info dictionary (if present)

    Raises:
        ValueError: If JSON is invalid or has unexpected structure
    """
    try:
        data = json.loads(content)
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON: {e}")

    # Handle both old format (just entries list) and new format (object with entries key)
    if isinstance(data, list):
        # Old format: just a list of entries
        return {"entries": data, "leave_days": [], "period": None}

    if not isinstance(data, dict):
        raise ValueError("JSON must be an object or array of entries")

    entries = data.get("entries", [])
    if not isinstance(entries, list):
        raise ValueError("'entries' field must be an array")

    leave_days = data.get("leave_days", [])
    if not isinstance(leave_days, list):
        leave_days = []

    period = data.get("period")

    return {
        "entries": entries,
        "leave_days": leave_days,
        "period": period,
    }


def parse_time_entries_from_csv(content: str) -> Dict:
    """
    Parse CSV content into structured import data.

    Handles CSV files with or without Category column.
    Skips summary statistics rows at the end.

    Args:
        content: CSV string content

    Returns:
        Dictionary with:
        - entries: List of entry dictionaries
        - leave_days: Empty list (CSV doesn't support leave days)

    Raises:
        ValueError: If CSV is empty or has no valid rows
    """
    reader = csv.DictReader(io.StringIO(content))

    entries = []
    for row in reader:
        # Skip empty rows and summary statistics rows
        if not row or not row.get("Date"):
            continue

        # Check if this is a summary row (no start time)
        if not row.get("Start Time"):
            # Likely hit the summary section, stop processing
            break

        entry = {
            "date": row.get("Date", "").strip(),
            "start_time": row.get("Start Time", "").strip(),
            "end_time": row.get("End Time", "").strip(),
            "description": row.get("Description", "").strip(),
        }

        # Optional Category column
        if "Category" in row and row["Category"]:
            entry["category"] = row["Category"].strip()

        # Optional duration (will be recalculated if missing)
        duration_str = row.get("Duration (Hours)", "").strip()
        if duration_str:
            try:
                entry["duration_hours"] = float(duration_str)
            except ValueError:
                pass

        # Optional created_at
        created_at = row.get("Created At", "").strip()
        if created_at:
            entry["created_at"] = created_at

        entries.append(entry)

    if not entries:
        raise ValueError("CSV file contains no valid time entries")

    return {
        "entries": entries,
        "leave_days": [],  # CSV doesn't support leave days
    }


def parse_time_entries_from_excel(content: bytes) -> Dict:
    """
    Parse Excel content into structured import data.

    Reads from the first sheet (Time Entries).

    Args:
        content: Excel file content as bytes

    Returns:
        Dictionary with:
        - entries: List of entry dictionaries
        - leave_days: Empty list (Excel doesn't support leave days yet)

    Raises:
        ValueError: If Excel file is invalid or has no valid rows
    """
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True)

    # Get the first sheet
    sheet = workbook.active

    # Read headers from first row
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value if cell.value else "")

    # Map headers to column indices
    header_map = {h: i for i, h in enumerate(headers)}

    # Required columns
    date_col = header_map.get("Date")
    start_col = header_map.get("Start Time")
    end_col = header_map.get("End Time")

    if date_col is None or start_col is None or end_col is None:
        raise ValueError(
            "Excel file must have 'Date', 'Start Time', and 'End Time' columns"
        )

    # Optional columns
    desc_col = header_map.get("Description")
    cat_col = header_map.get("Category")
    duration_col = header_map.get("Duration (Hours)")
    created_col = header_map.get("Created At")

    entries = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        # Skip empty rows
        if not row or not row[date_col]:
            continue

        # Format date - handle both string and date objects
        date_val = row[date_col]
        if isinstance(date_val, date):
            date_str = date_val.isoformat()
        elif date_val:
            date_str = str(date_val)
        else:
            continue

        # Format times - handle both string and time objects
        start_val = row[start_col]
        if isinstance(start_val, datetime_time):
            start_str = start_val.strftime("%H:%M:%S")
        elif start_val:
            start_str = str(start_val)
        else:
            continue

        end_val = row[end_col]
        if isinstance(end_val, datetime_time):
            end_str = end_val.strftime("%H:%M:%S")
        elif end_val:
            end_str = str(end_val)
        else:
            continue

        entry = {
            "date": date_str,
            "start_time": start_str,
            "end_time": end_str,
        }

        # Optional fields
        if desc_col is not None and row[desc_col]:
            entry["description"] = str(row[desc_col])

        if cat_col is not None and row[cat_col]:
            entry["category"] = str(row[cat_col])

        if duration_col is not None and row[duration_col]:
            try:
                entry["duration_hours"] = float(row[duration_col])
            except (ValueError, TypeError):
                pass

        if created_col is not None and row[created_col]:
            created_val = row[created_col]
            if isinstance(created_val, datetime):
                entry["created_at"] = created_val.isoformat()
            elif created_val:
                entry["created_at"] = str(created_val)

        entries.append(entry)

    workbook.close()

    if not entries:
        raise ValueError("Excel file contains no valid time entries")

    return {
        "entries": entries,
        "leave_days": [],  # Excel doesn't support leave days yet
    }


def validate_leave_day_data(leave: Dict) -> Tuple[bool, List[str]]:
    """
    Validate a single leave day dictionary.

    Args:
        leave: Dictionary with leave day data

    Returns:
        Tuple of (is_valid, list_of_error_messages)
    """
    errors = []

    # Required fields
    if not leave.get("date"):
        errors.append("Missing required field: date")
    else:
        parsed_date = normalize_date_string(str(leave["date"]))
        if parsed_date is None:
            errors.append(f"Invalid date format: {leave['date']}")

    leave_type = leave.get("leave_type", "").lower() if leave.get("leave_type") else ""
    if not leave_type:
        errors.append("Missing required field: leave_type")
    elif leave_type not in ("vacation", "sick"):
        errors.append(f"Invalid leave_type: {leave_type}. Must be 'vacation' or 'sick'")

    return len(errors) == 0, errors
