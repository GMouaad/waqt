"""Unit tests for JSON and Excel export functionality."""

import pytest
import json
import io
import openpyxl
from datetime import date, time
from click.testing import CliRunner

@pytest.fixture
def app():
    """Create and configure a test app instance."""
    from src.waqt import create_app, db
    
    app = create_app()
    app.config["TESTING"] = True
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///:memory:"

    with app.app_context():
        db.create_all()
        yield app
        db.session.remove()
        db.drop_all()


@pytest.fixture
def client(app):
    """Create a test client for the app."""
    return app.test_client()


@pytest.fixture
def cli():
    """Return the CLI entry point."""
    from src.waqt.cli import cli as cli_obj
    return cli_obj


@pytest.fixture
def runner():
    """Create a CLI test runner."""
    return CliRunner()


@pytest.fixture
def sample_entries(app):
    """Create sample time entries for testing."""
    from src.waqt.models import TimeEntry, Category
    from src.waqt import db
    
    with app.app_context():
        category = Category(name="Work", color="#ff0000")
        db.session.add(category)
        db.session.commit()
        
        entries = [
            TimeEntry(
                date=date(2024, 1, 15),
                start_time=time(9, 0),
                end_time=time(17, 0),
                duration_hours=8.0,
                description="Regular work day",
                category_id=category.id
            ),
            TimeEntry(
                date=date(2024, 1, 16),
                start_time=time(9, 0),
                end_time=time(18, 0),
                duration_hours=9.0,
                description="Overtime day",
            ),
        ]
        for entry in entries:
            db.session.add(entry)
        db.session.commit()
        return entries


def test_export_json_structure(app, sample_entries):
    """Test JSON export structure and content."""
    from src.waqt.utils import export_time_entries_to_json
    from src.waqt.models import TimeEntry
    
    with app.app_context():
        entries = TimeEntry.query.all()
        json_content = export_time_entries_to_json(entries)
        
        data = json.loads(json_content)
        
        assert "period" in data
        assert "entries" in data
        assert "summary" in data
        
        assert len(data["entries"]) == 2
        assert data["entries"][0]["date"] == "2024-01-15"
        assert data["entries"][0]["category"] == "Work"
        assert data["entries"][1]["overtime"] == 1.0
        
        assert data["summary"]["total_entries"] == 2
        assert data["summary"]["total_hours"] == 17.0
        assert data["summary"]["total_overtime"] == 1.0


def test_export_excel_structure(app, sample_entries):
    """Test Excel export structure and content."""
    from src.waqt.utils import export_time_entries_to_excel
    from src.waqt.models import TimeEntry
    
    with app.app_context():
        entries = TimeEntry.query.all()
        excel_content = export_time_entries_to_excel(entries)
        
        # Load the Excel file
        workbook = openpyxl.load_workbook(io.BytesIO(excel_content))
        
        assert "Time Entries" in workbook.sheetnames
        assert "Summary" in workbook.sheetnames
        
        sheet = workbook["Time Entries"]
        
        # Check headers
        headers = [cell.value for cell in sheet[1]]
        assert "Date" in headers
        assert "Start Time" in headers
        assert "Category" in headers
        
        # Check data
        # Row 2 (first entry)
        # OpenPyXL returns datetime objects for dates
        cell_value = sheet.cell(row=2, column=1).value
        # Handle both datetime and date objects
        val_date = cell_value.date() if hasattr(cell_value, "date") else cell_value
        assert val_date == date(2024, 1, 15)
        
        assert sheet.cell(row=2, column=8).value == "Work" # Category column
        
        # Row 3 (second entry) - Check overtime
        # Column 9 is Overtime based on utils.py implementation
        assert sheet.cell(row=3, column=9).value == 1.0


def test_export_json_route(client, app, sample_entries):
    """Test JSON export via Flask route."""
    with app.app_context():
        response = client.get("/export/json?period=all")
        
        assert response.status_code == 200
        assert response.mimetype == "application/json"
        
        data = json.loads(response.data)
        assert len(data["entries"]) == 2


def test_export_excel_route(client, app, sample_entries):
    """Test Excel export via Flask route."""
    with app.app_context():
        response = client.get("/export/excel?period=all")
        
        assert response.status_code == 200
        assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        # Check if it's a valid zip file (xlsx is a zip)
        assert response.data.startswith(b"PK")


def test_cli_export_json(runner, app, cli, sample_entries, tmp_path):
    """Test CLI export to JSON."""
    with app.app_context():
        output_file = tmp_path / "export.json"
        
        result = runner.invoke(cli, ["export", "--format", "json", "--output", str(output_file)])
        
        assert result.exit_code == 0
        assert output_file.exists()
        
        with open(output_file) as f:
            data = json.load(f)
            assert len(data["entries"]) == 2


def test_cli_export_excel(runner, app, cli, sample_entries, tmp_path):
    """Test CLI export to Excel."""
    with app.app_context():
        output_file = tmp_path / "export.xlsx"
        
        result = runner.invoke(cli, ["export", "--format", "excel", "--output", str(output_file)])
        
        assert result.exit_code == 0
        assert output_file.exists()
        
        # Verify it's a valid excel file
        workbook = openpyxl.load_workbook(output_file)
        assert "Time Entries" in workbook.sheetnames
